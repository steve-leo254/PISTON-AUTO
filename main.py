import os
from functools import wraps
from datetime import datetime, timedelta

from werkzeug.security import generate_password_hash, check_password_hash
from flask import Flask, render_template, request, redirect, url_for, flash, session, jsonify
from flask_sqlalchemy import SQLAlchemy
from flask_wtf import FlaskForm
from wtforms import StringField, PasswordField, SubmitField, StringField, FloatField, IntegerField, SelectField, SubmitField
from wtforms.validators import DataRequired, Length, Email, EqualTo, NumberRange
from enum import Enum
from flask_bcrypt import Bcrypt
from flask_migrate import Migrate
from sqlalchemy.exc import IntegrityError
from sqlalchemy.sql import func

from dbservice import (
    create_app,
    db,
    User,
    UserRole,
    Product,
    ProductForm,
    Sale,
    SaleForm,
    LoginForm,
    RegisterForm,
    Booking,
    BookingForm,
    BookingStatus,
    JobPartUsage,
    Invoice,
    InvoiceStatus,
    InvoiceItem,
)


app = create_app()

app.secret_key = os.environ.get('SECRET_KEY', 'default_secret_key')


def login_required(role: UserRole | None = None):
    def decorator(view_func):
        @wraps(view_func)
        def wrapped(*args, **kwargs):
            username = session.get('username')
            if not username:
                flash('Please log in to continue.', 'warning')
                return redirect(url_for('login'))

            user = User.query.filter_by(username=username).first()
            if not user:
                session.pop('username', None)
                flash('Account not found. Please log in again.', 'warning')
                return redirect(url_for('login'))

            if role is not None and user.role != role:
                flash('Access denied. Insufficient permissions.', 'danger')
                return redirect(url_for('login'))

            return view_func(*args, **kwargs)

        return wrapped

    return decorator


@app.route('/')
def landing():
    # Check if any admin users exist
    admin_count = User.query.filter_by(role=UserRole.ADMIN).count()
    return render_template('index.html', admin_count=admin_count)



@app.route('/register', methods=['GET', 'POST'])
def register():
    # Allow regular user registration regardless of admin existence
    # Only prevent setup-admin if admin already exists
    form = RegisterForm()

    if form.validate_on_submit():
        username = form.username.data
        email = form.email.data
        password = form.password.data

        hashed_password = generate_password_hash(password)

        new_user = User(
            username=username,
            email=email,
            password=hashed_password,
            role=UserRole.USER
        )

        try:
            db.session.add(new_user)
            db.session.commit()
            flash('Registration successful! You can now login.', 'success')
            return redirect(url_for('login'))

        except IntegrityError as e:
            # Rollback the session to avoid leaving the database in an inconsistent state
            db.session.rollback()

            # Check if the error is due to a duplicate username
            if 'duplicate key value violates unique constraint "user_username_key"' in str(e):
                flash(
                    'Username already exists. Please choose a different one.', 'danger')
            else:
                app.logger.error(f'Error during registration: {str(e)}')
                flash(
                    'An error occurred during registration. Please try again.', 'danger')

    return render_template('register.html', form=form)

# ... (rest of the code)
@app.route('/login', methods=['GET', 'POST'])
def login():
    print("DEBUG: Login route accessed")
    form = LoginForm()
    
    print(f"DEBUG: Form method: {request.method}")
    print(f"DEBUG: Form validate_on_submit: {form.validate_on_submit()}")
    if form.errors:
        print(f"DEBUG: Form errors: {form.errors}")

    if form.validate_on_submit():
        email = form.email.data
        password = form.password.data
        print(f"DEBUG: Email: {email}")
        print(f"DEBUG: Password provided: {'Yes' if password else 'No'}")

        user = User.query.filter_by(email=email).first()
        print(f"DEBUG: User found: {'Yes' if user else 'No'}")

        if user and check_password_hash(user.password, password):
            # Store username and role in the session
            session['username'] = user.username
            # Convert enum to string for session
            session['user_role'] = user.role.value
            
            # Debug: Print session data
            print(f"DEBUG: Login successful - Username: {user.username}, Role: {user.role.value}")
            print(f"DEBUG: Session data: {dict(session)}")

            flash('Login successful!', 'success')
            if user.role == UserRole.ADMIN:
                return redirect(url_for('admin_dashboard'))
            elif user.role == UserRole.USER:
                return redirect(url_for('user_dashboard'))
        else:
            flash('Invalid email or password. Please try again.', 'danger')

    return render_template('login.html', form=form)


@app.route('/admin-dashboard')
@login_required(UserRole.ADMIN)
def admin_dashboard():
    username = session['username']
    
    # Basic metrics for quick overview
    total_users = User.query.count()
    total_bookings = Booking.query.count()
    total_products = Product.query.count()

    # Sales calculations
    total_sales_amount = (
        db.session.query(
            func.coalesce(
                func.sum(Sale.quantity * Product.selling_price), 0
            )
        )
        .join(Product, Product.id == Sale.product_id)
        .scalar()
    )
    
    # Enhanced metrics
    total_cost = (
        db.session.query(
            func.coalesce(
                func.sum(Sale.quantity * Product.buying_price), 0
            )
        )
        .join(Product, Product.id == Sale.product_id)
        .scalar()
    )
    
    gross_profit = total_sales_amount - total_cost
    
    # Recent bookings (last 7 days)
    recent_bookings = (
        Booking.query.filter(Booking.service_date >= datetime.now() - timedelta(days=7))
        .order_by(Booking.service_date.desc())
        .limit(5)
        .all()
    )
    
    # Booking status breakdown
    booking_status_counts = (
        db.session.query(
            Booking.status,
            func.count(Booking.id).label('count')
        )
        .group_by(Booking.status)
        .all()
    )
    
    # Convert to dictionary with string keys for JSON serialization
    booking_status_counts = {status.value: count for status, count in booking_status_counts}
    print(f"DEBUG: booking_status_counts type: {type(booking_status_counts)}")
    print(f"DEBUG: booking_status_counts keys: {list(booking_status_counts.keys())}")
    print(f"DEBUG: booking_status_counts: {booking_status_counts}")
    
    # Recent sales (last 7 days)
    recent_sales = (
        db.session.query(Sale, Product)
        .join(Product, Product.id == Sale.product_id)
        .filter(Sale.created_at >= datetime.now() - timedelta(days=7))
        .order_by(Sale.created_at.desc())
        .limit(5)
        .all()
    )
    
    # Low stock products
    low_stock_products = (
        db.session.query(
            Product,
            func.coalesce(
                db.session.query(func.sum(JobPartUsage.quantity_used))
                .filter(JobPartUsage.product_id == Product.id)
                .correlate(Product)
                .scalar_subquery(), 0
            ).label('total_used')
        )
        .filter(Product.stock_quantity <= 20)
        .order_by(Product.stock_quantity.asc())
        .limit(5)
        .all()
    )
    
    # Today's activity
    today_bookings = Booking.query.filter(
        func.date(Booking.service_date) == datetime.now().date()
    ).count()
    
    today_sales = (
        db.session.query(func.coalesce(func.sum(Sale.quantity * Product.selling_price), 0))
        .join(Product, Product.id == Sale.product_id)
        .filter(func.date(Sale.created_at) == datetime.now().date())
        .scalar()
    )
    
    # Monthly trends (last 6 months)
    monthly_revenue = []
    monthly_labels = []
    for i in range(6):
        month_start = datetime.now().replace(day=1) - timedelta(days=30*i)
        month_end = (month_start + timedelta(days=32)).replace(day=1) - timedelta(days=1)
        
        revenue = (
            db.session.query(func.coalesce(func.sum(Sale.quantity * Product.selling_price), 0))
            .join(Product, Product.id == Sale.product_id)
            .filter(Sale.created_at >= month_start, Sale.created_at <= month_end)
            .scalar()
        )
        
        monthly_revenue.append(float(revenue))
        monthly_labels.append(month_start.strftime('%b'))

    return render_template(
        'admin_dashboard.html',
        username=username,
        total_users=total_users,
        total_bookings=total_bookings,
        total_products=total_products,
        total_sales_amount=total_sales_amount,
        total_cost=total_cost,
        gross_profit=gross_profit,
        recent_bookings=recent_bookings,
        recent_sales=recent_sales,
        low_stock_products=low_stock_products,
        booking_status_counts=booking_status_counts,
        today_bookings=today_bookings,
        today_sales=today_sales,
        monthly_revenue=monthly_revenue[::-1],  # Reverse to show oldest to newest
        monthly_labels=monthly_labels[::-1],
        datetime=datetime,  # Make datetime available in template
    )


@app.route('/user-dashboard')
@login_required(UserRole.USER)
def user_dashboard():
    username = session['username']
    user = User.query.filter_by(username=username).first()

    upcoming_bookings = (
        Booking.query.filter_by(user_id=user.id)
        .order_by(Booking.service_date.desc())
        .limit(5)
        .all()
    )

    # Convert bookings to JSON-serializable format
    bookings_data = {}
    for booking in upcoming_bookings:
        booking_dict = {
            'id': booking.id,
            'status': booking.status.value,
            'service_date': booking.service_date.isoformat() if booking.service_date else None,
            'service_type': booking.service_type,
            'vehicle_type': booking.vehicle_type,
            'customer_name': booking.customer_name,
            'email': booking.email,
            'phone': booking.phone,
            'special_request': booking.special_request
        }
        bookings_data[str(booking.id)] = booking_dict

    return render_template(
        'user_dashboard.html', username=username, bookings=bookings_data
    )


@app.route('/products')
@login_required(UserRole.ADMIN)
def products():
    form = ProductForm()
    products = Product.query.all()
    return render_template('products.html', products=products, form=form)


@app.route('/sales')
@login_required(UserRole.ADMIN)
def sales():
    form = SaleForm()
    sales = Sale.query.all()
    return render_template('sales.html', sales=sales, form=form)




@app.route('/add_sale', methods=['GET', 'POST'])
@login_required(UserRole.ADMIN)
def add_sale():
    form = SaleForm()

    # Fetch products to populate the choices in the SaleForm
    form.product_id.choices = [(product.id, product.name)
                               for product in Product.query.all()]

    if form.validate_on_submit():
        product_id = form.product_id.data
        quantity = form.quantity.data

        product = Product.query.get(product_id)

        if product and product.stock_quantity >= quantity:
            # Reduce stock quantity of the product
            product.stock_quantity -= quantity

            try:
                new_sale = Sale(product_id=product_id, quantity=quantity)
                db.session.add(new_sale)
                db.session.commit()

                flash(
                    f'Sale recorded successfully for {product.name}!', 'success')
            except Exception as e:
                db.session.rollback()
                app.logger.error(f'Error recording sale: {str(e)}')
                flash('Error recording sale. Please try again.', 'danger')
        else:
            flash('Invalid product or insufficient stock quantity.', 'danger')

        return redirect(url_for('sales'))

    return render_template('sales.html', form=form)


@app.route('/book-service', methods=['GET', 'POST'])
@login_required(UserRole.USER)
def book_service():
    form = BookingForm()
    
    print(f"DEBUG: Form method: {request.method}")
    print(f"DEBUG: Form validate_on_submit: {form.validate_on_submit()}")
    if form.errors:
        print(f"DEBUG: Form errors: {form.errors}")

    if form.validate_on_submit():
        username = session['username']
        user = User.query.filter_by(username=username).first()

        service_date_str = form.service_date.data
        try:
            service_date = datetime.strptime(service_date_str, '%Y-%m-%d')
        except ValueError:
            flash('❌ Invalid service date format. Please use YYYY-MM-DD format.', 'danger')
            return render_template('booking.html', form=form)
        
        # Validate that the date is in the future
        if service_date.date() < datetime.now().date():
            flash('❌ Service date cannot be in the past. Please select a future date.', 'danger')
            return render_template('booking.html', form=form)

        booking = Booking(
            user_id=user.id,
            customer_name=form.customer_name.data,
            email=form.email.data,
            phone=form.phone.data,
            vehicle_type=form.vehicle_type.data,
            service_type=form.service_type.data,
            service_date=service_date,
            special_request=form.special_request.data
        )

        try:
            print(f"DEBUG: About to add booking to database")
            print(f"DEBUG: Booking data: {booking}")
            db.session.add(booking)
            print(f"DEBUG: Booking added to session")
            db.session.commit()
            print(f"DEBUG: Database commit successful")

            # Success notifications
            flash('🎉 Booking created successfully!', 'success')
            flash(f'📧 Confirmation email will be sent to: {form.email.data}', 'info')
            flash(f'🔧 Service Type: {form.service_type.data}', 'info')
            flash(f'📅 Service Date: {service_date.strftime("%B %d, %Y")}', 'info')
            flash('📋 We will contact you soon to confirm your booking.', 'warning')
            
            # Store booking details for success page
            session['booking_success'] = {
                'customer_name': form.customer_name.data,
                'service_type': form.service_type.data,
                'service_date': service_date.strftime("%B %d, %Y"),
                'email': form.email.data,
                'phone': form.phone.data,
                'vehicle_type': form.vehicle_type.data
            }
            
            return redirect(url_for('my_bookings'))
            
        except Exception as e:
            print(f"DEBUG: Exception occurred: {str(e)}")
            print(f"DEBUG: Exception type: {type(e)}")
            db.session.rollback()
            app.logger.error(f'Error creating booking: {str(e)}')
            flash('❌ Error creating booking. Please try again.', 'danger')
            flash('If the problem persists, please contact our support team.', 'warning')

    return render_template('booking.html', form=form)


    
@app.route('/booking-success')
@login_required(UserRole.USER)
def booking_success():
    booking_details = session.pop('booking_success', None)
    
    if not booking_details:
        return redirect(url_for('my_bookings'))
    
    return render_template('booking_success.html', booking=booking_details)


@app.route('/about')
def about():
    return render_template('about.html')


@app.route('/services')
def services():
    # Define services data for the template
    services = [
        {
            'id': 1,
            'name': 'Engine Diagnostics',
            'description': 'Comprehensive engine diagnostics using state-of-the-art equipment to identify and resolve any engine issues quickly and accurately.',
            'icon': 'fa-cogs',
            'image': '/img/service-1.jpg',
            'features': ['Complete engine scan', 'Performance analysis', 'Emission testing', 'Computer diagnostics'],
            'time_required': '1-2 hours',
            'price': 'Ksh 3,500 - 8,000'
        },
        {
            'id': 2,
            'name': 'Oil Change & Service',
            'description': 'Regular oil changes and comprehensive service to keep your engine running smoothly and efficiently.',
            'icon': 'fa-oil-can',
            'image': '/img/service-2.jpg',
            'features': ['Oil filter replacement', 'Fluid check', 'Lubrication', 'Safety inspection'],
            'time_required': '30-45 minutes',
            'price': 'Ksh 2,500 - 4,000'
        },
        {
            'id': 3,
            'name': 'Brake Service',
            'description': 'Complete brake system inspection and service to ensure your safety on the road.',
            'icon': 'fa-compact-disc',
            'image': '/img/service-3.jpg',
            'features': ['Brake pad replacement', 'Rotor resurfacing', 'Fluid flush', 'System inspection'],
            'time_required': '2-3 hours',
            'price': 'Ksh 5,000 - 12,000'
        },
        {
            'id': 4,
            'name': 'Tire Service',
            'description': 'Complete tire care including rotation, balancing, and replacement for optimal performance and safety.',
            'icon': 'fa-circle',
            'image': '/img/service-4.jpg',
            'features': ['Tire rotation', 'Wheel balancing', 'Patch repair', 'New tire installation'],
            'time_required': '1-2 hours',
            'price': 'Ksh 1,500 - 15,000'
        }
    ]
    return render_template('services.html', services=services)


@app.route('/contact')
def contact():
    return render_template('contact.html')


@app.route('/testimonial')
def testimonial():
    return render_template('testimonial.html')


@app.route('/404')
def not_found():
    return render_template('404.html'), 404


@app.route('/bookings')
@login_required(UserRole.USER)
def my_bookings():
    username = session['username']
    print(f"DEBUG: My Bookings - Username: {username}")
    user = User.query.filter_by(username=username).first()
    print(f"DEBUG: My Bookings - User found: {user.id if user else 'None'}")
    
    bookings_query = (
        Booking.query.filter_by(user_id=user.id)
        .order_by(Booking.service_date.desc())
    ) if user else None
    
    bookings = {}
    if bookings_query:
        for booking in bookings_query.all():
            booking_dict = {
                'id': booking.id,
                'user_id': booking.user_id,
                'customer_name': booking.customer_name,
                'email': booking.email,
                'phone': booking.phone,
                'vehicle_type': booking.vehicle_type,
                'service_type': booking.service_type,
                'service_date': booking.service_date.isoformat() if booking.service_date else None,
                'special_request': booking.special_request,
                'status': booking.status.value if booking.status else None,
                'estimated_total': booking.estimated_total
            }
            bookings[str(booking.id)] = booking_dict
    
    print(f"DEBUG: My Bookings - Found {len(bookings)} bookings")
    for booking_id, booking in bookings.items():
        print(f"DEBUG: Booking ID: {booking['id']}, Service: {booking['service_type']}, Date: {booking['service_date']}")
    
    return render_template('my_bookings.html', bookings=bookings)


@app.route('/admin/bookings')
@login_required(UserRole.ADMIN)
def admin_bookings():
    bookings = (
        Booking.query.order_by(Booking.service_date.desc())
        .join(User, Booking.user_id == User.id)
        .add_columns(User.username)
        .all()
    )
    return render_template('admin_bookings.html', bookings=bookings, timedelta=timedelta, datetime=datetime)


@app.route('/admin/bookings/update/<int:booking_id>', methods=['POST'])
@login_required(UserRole.ADMIN)
def update_booking_status(booking_id):
    """Update booking status"""
    
    booking = Booking.query.get_or_404(booking_id)
    
    # Get new status from request
    data = request.get_json()
    new_status = data.get('status') if data else None
    
    if not new_status:
        return jsonify({'success': False, 'message': 'Status not provided'}), 400
    
    # Validate status
    try:
        booking.status = BookingStatus[new_status]
        db.session.commit()
        return jsonify({
            'success': True, 
            'message': f'Booking status updated to {new_status}',
            'new_status': new_status
        })
    except (KeyError, ValueError) as e:
        return jsonify({'success': False, 'message': f'Invalid status: {new_status}'}), 400
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/admin/users')
@login_required(UserRole.ADMIN)
def admin_users():
    """Admin users page - manage system users"""
    users = User.query.order_by(User.id.desc()).all()
    return render_template('admin_users.html', users=users)


@app.route('/admin/sales')
@login_required(UserRole.ADMIN)
def admin_sales():
    """Admin sales page - view sales analytics"""
    # Get sales data with product info
    sales_data = (
        db.session.query(
            Sale.id,
            Sale.quantity,
            Sale.created_at,
            Product.name,
            Product.buying_price,
            Product.selling_price,
            func.coalesce(Sale.quantity * Product.selling_price, 0).label('revenue'),
            func.coalesce(Sale.quantity * Product.buying_price, 0).label('cost'),
            func.coalesce(Sale.quantity * (Product.selling_price - Product.buying_price), 0).label('profit')
        )
        .join(Product, Product.id == Sale.product_id)
        .order_by(Sale.created_at.desc())
        .all()
    )
    
    # Calculate totals
    total_revenue = sum(sale.revenue for sale in sales_data)
    total_cost = sum(sale.cost for sale in sales_data)
    total_profit = sum(sale.profit for sale in sales_data)
    
    return render_template('admin_sales.html', 
                       sales_data=sales_data,
                       total_revenue=total_revenue,
                       total_cost=total_cost,
                       total_profit=total_profit)


@app.route('/admin/bookings/export')
@login_required(UserRole.ADMIN)
def export_bookings():
    """Export bookings as Excel file"""
    import pandas as pd
    from io import BytesIO
    from flask import Response
    
    # Get all bookings with user info
    bookings = (
        Booking.query.order_by(Booking.service_date.desc())
        .join(User, Booking.user_id == User.id)
        .add_columns(User.username)
        .all()
    )
    
    # Prepare data for DataFrame
    booking_data = []
    for booking, username in bookings:
        booking_data.append({
            'Booking ID': booking.id,
            'Customer Name': booking.customer_name,
            'Email': booking.email,
            'Phone': booking.phone,
            'Vehicle Type': booking.vehicle_type,
            'Service Type': booking.service_type,
            'Service Date': booking.service_date.strftime('%Y-%m-%d %H:%M') if booking.service_date else '',
            'Status': booking.status.value.replace('_', ' ').title(),
            'Created Date': booking.created_at.strftime('%Y-%m-%d %H:%M'),
            'Created By': username,
            'Special Request': booking.special_request or ''
        })
    
    # Create DataFrame
    df = pd.DataFrame(booking_data)
    
    # Create Excel file in memory
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Write main data
        df.to_excel(writer, sheet_name='Bookings', index=False)
        
        # Get the workbook and worksheet for formatting
        workbook = writer.book
        worksheet = writer.sheets['Bookings']
        
        # Format headers
        from openpyxl.styles import Font, PatternFill, Border, Side
        
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                            top=Side(style='thin'), bottom=Side(style='thin'))
        
        # Apply formatting to header row
        for col_num, header in enumerate(df.columns, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.style = header_font
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    # Create response
    output.seek(0)
    response = Response(
        output.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={
            'Content-Disposition': f'attachment; filename=bookings_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        }
    )
    
    return response


@app.route('/admin/stock')
@login_required(UserRole.ADMIN)
def admin_stock():
    products = Product.query.all()

    usage_subquery = (
        db.session.query(
            JobPartUsage.product_id,
            func.coalesce(func.sum(JobPartUsage.quantity_used), 0).label(
                'total_used'
            )
        )
        .group_by(JobPartUsage.product_id)
        .subquery()
    )

    stock_rows = (
        db.session.query(
            Product,
            func.coalesce(
                db.session.query(func.sum(JobPartUsage.quantity_used))
                .filter(JobPartUsage.product_id == Product.id)
                .correlate(Product)
                .scalar_subquery(), 0
            ).label('total_used')
        )
        .outerjoin(
            usage_subquery, Product.id == usage_subquery.c.product_id
        )
        .all()
    )

    return render_template('admin_stock.html', stock_rows=stock_rows)


@app.route('/admin/stock/export')
@login_required(UserRole.ADMIN)
def export_stock():
    """Export stock as Excel file"""
    import pandas as pd
    from io import BytesIO
    from flask import Response
    
    # Get all products with usage data
    products = Product.query.all()
    
    usage_subquery = (
        db.session.query(
            JobPartUsage.product_id,
            func.coalesce(func.sum(JobPartUsage.quantity_used), 0).label(
                'total_used'
            )
        )
        .group_by(JobPartUsage.product_id)
        .subquery()
    )

    stock_data = []
    for product in products:
        # Get usage for this product
        usage_info = (
            db.session.query(usage_subquery.c.total_used)
            .filter(usage_subquery.c.product_id == product.id)
            .scalar()
        ) or 0
        
        available = product.stock_quantity - usage_info
        profit_margin = ((product.selling_price - product.buying_price) / product.buying_price * 100) if product.buying_price > 0 else 0
        
        stock_data.append({
            'Product ID': product.id,
            'Product Name': product.name,
            'Description': product.description or '',
            'Buying Price': product.buying_price,
            'Selling Price': product.selling_price,
            'Profit Margin (%)': round(profit_margin, 2),
            'Stock Quantity': product.stock_quantity,
            'Total Used': usage_info,
            'Available': available,
            'Stock Status': 'Critical' if available <= 5 else 'Low' if available <= 20 else 'Medium' if available <= 50 else 'Good',
            'Total Value': product.selling_price * available
        })
    
    # Create DataFrame
    df = pd.DataFrame(stock_data)
    
    # Create Excel file in memory
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Write main data
        df.to_excel(writer, sheet_name='Stock Report', index=False)
        
        # Get workbook and worksheet for formatting
        workbook = writer.book
        worksheet = writer.sheets['Stock Report']
        
        # Format headers
        from openpyxl.styles import Font, PatternFill, Border, Side
        
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                            top=Side(style='thin'), bottom=Side(style='thin'))
        
        # Format numbers
        from openpyxl.styles import NamedStyle
        
        number_format = NamedStyle(name="number_format")
        number_format.number_format = '#,##0.00'
        
        # Apply formatting to header row
        for col_num, header in enumerate(df.columns, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.style = header_font
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    # Create response
    output.seek(0)
    response = Response(
        output.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={
            'Content-Disposition': f'attachment; filename=stock_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        }
    )
    
    return response


@app.route('/admin/stock/add', methods=['GET', 'POST'])
@login_required(UserRole.ADMIN)
def add_product():
    """Add new product"""
    if request.method == 'POST':
        try:
            # Get form data
            name = request.form.get('name')
            category = request.form.get('category')
            description = request.form.get('description')
            buying_price = float(request.form.get('buying_price', 0))
            stock_quantity = int(request.form.get('stock_quantity', 0))
            
            # Validate required fields
            if not name or buying_price <= 0:
                flash('Product name and purchase price are required!', 'danger')
                return redirect(url_for('admin_stock'))
            
            # Create new product
            new_product = Product(
                name=name,
                description=description,
                buying_price=buying_price,
                selling_price=buying_price * 1.3,  # Default markup for internal tracking
                stock_quantity=stock_quantity
            )
            
            db.session.add(new_product)
            db.session.commit()
            
            flash(f'Product "{name}" added successfully!', 'success')
            return redirect(url_for('admin_stock'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error adding product: {str(e)}', 'danger')
            return redirect(url_for('admin_stock'))
    
    return render_template('add_product.html')


@app.route('/admin/stock/edit/<int:product_id>', methods=['GET', 'POST'])
@login_required(UserRole.ADMIN)
def edit_product(product_id):
    """Edit existing product"""
    product = Product.query.get_or_404(product_id)
    
    if request.method == 'POST':
        try:
            # Get form data
            name = request.form.get('name')
            category = request.form.get('category')
            description = request.form.get('description')
            buying_price = float(request.form.get('buying_price', 0))
            selling_price = float(request.form.get('selling_price', 0))
            stock_quantity = int(request.form.get('stock_quantity', 0))
            
            # Update product
            product.name = name
            if hasattr(product, 'category'):
                product.category = category
            product.description = description
            product.buying_price = buying_price
            product.selling_price = selling_price
            product.stock_quantity = stock_quantity
            
            db.session.commit()
            
            flash(f'Product "{name}" updated successfully!', 'success')
            return redirect(url_for('admin_stock'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error updating product: {str(e)}', 'danger')
            return redirect(url_for('admin_stock'))
    
    return render_template('edit_product.html', product=product)


@app.route('/admin/financial-reports')
@login_required(UserRole.ADMIN)
def admin_financial_reports():
    # Current month data for comparison
    current_month = datetime.now().month
    current_year = datetime.now().year
    
    # Previous month data
    if current_month == 1:
        prev_month = 12
        prev_year = current_year - 1
    else:
        prev_month = current_month - 1
        prev_year = current_year
    
    # Total revenue (all time) - for main display
    total_revenue_all = (
        db.session.query(
            func.coalesce(
                func.sum(Sale.quantity * Product.selling_price), 0
            )
        )
        .join(Product, Product.id == Sale.product_id)
        .scalar()
    )
    
    # Total revenue (current month) - for comparison
    total_revenue = (
        db.session.query(
            func.coalesce(
                func.sum(Sale.quantity * Product.selling_price), 0
            )
        )
        .join(Product, Product.id == Sale.product_id)
        .filter(func.extract('month', Sale.created_at) == current_month,
                func.extract('year', Sale.created_at) == current_year)
        .scalar()
    )
    
    # Previous month revenue
    prev_revenue = (
        db.session.query(
            func.coalesce(
                func.sum(Sale.quantity * Product.selling_price), 0
            )
        )
        .join(Product, Product.id == Sale.product_id)
        .filter(func.extract('month', Sale.created_at) == prev_month,
                func.extract('year', Sale.created_at) == prev_year)
        .scalar()
    )
    
    # Total cost (all time)
    total_cost_all = (
        db.session.query(
            func.coalesce(
                func.sum(Sale.quantity * Product.buying_price), 0
            )
        )
        .join(Product, Product.id == Sale.product_id)
        .scalar()
    )
    
    # Total cost (current month)
    total_cost = (
        db.session.query(
            func.coalesce(
                func.sum(Sale.quantity * Product.buying_price), 0
            )
        )
        .join(Product, Product.id == Sale.product_id)
        .filter(func.extract('month', Sale.created_at) == current_month,
                func.extract('year', Sale.created_at) == current_year)
        .scalar()
    )
    
    # Previous month cost
    prev_cost = (
        db.session.query(
            func.coalesce(
                func.sum(Sale.quantity * Product.buying_price), 0
            )
        )
        .join(Product, Product.id == Sale.product_id)
        .filter(func.extract('month', Sale.created_at) == prev_month,
                func.extract('year', Sale.created_at) == prev_year)
        .scalar()
    )
    
    # Use all-time data if current month is empty
    if total_revenue == 0:
        total_revenue = total_revenue_all
    if total_cost == 0:
        total_cost = total_cost_all
    
    gross_profit = total_revenue - total_cost
    prev_gross_profit = prev_revenue - prev_cost
    
    # Calculate percentage changes
    revenue_change = ((total_revenue - prev_revenue) / prev_revenue * 100) if prev_revenue > 0 else 0
    cost_change = ((total_cost - prev_cost) / prev_cost * 100) if prev_cost > 0 else 0
    profit_change = ((gross_profit - prev_gross_profit) / prev_gross_profit * 100) if prev_gross_profit > 0 else 0

    latest_sales = (
        db.session.query(Sale, Product)
        .join(Product, Product.id == Sale.product_id)
        .order_by(Sale.created_at.desc())
        .limit(20)
        .all()
    )

    return render_template(
        'admin_financial_reports.html',
        total_revenue=total_revenue,
        total_cost=total_cost,
        gross_profit=gross_profit,
        revenue_change=revenue_change,
        cost_change=cost_change,
        profit_change=profit_change,
        latest_sales=latest_sales,
    )


@app.route('/admin/reports/export')
@login_required(UserRole.ADMIN)
def export_financial_reports():
    """Export financial reports as Excel file"""
    import pandas as pd
    from io import BytesIO
    from flask import Response
    
    # Get comprehensive financial data
    sales_data = (
        db.session.query(
            Sale.id,
            Sale.quantity,
            Sale.created_at,
            Product.name,
            Product.buying_price,
            Product.selling_price,
            func.coalesce(Sale.quantity * Product.selling_price, 0).label('revenue'),
            func.coalesce(Sale.quantity * Product.buying_price, 0).label('cost'),
            func.coalesce(Sale.quantity * (Product.selling_price - Product.buying_price), 0).label('profit')
        )
        .join(Product, Product.id == Sale.product_id)
        .order_by(Sale.created_at.desc())
        .all()
    )
    
    # Convert to list of dictionaries
    data = []
    for sale in sales_data:
        data.append({
            'Sale ID': sale[0],
            'Product Name': sale[3],
            'Quantity Sold': sale[1],
            'Buying Price (KES)': sale[4],
            'Selling Price (KES)': sale[5],
            'Revenue (KES)': sale[6],
            'Cost (KES)': sale[7],
            'Profit (KES)': sale[8],
            'Sale Date': sale[2].strftime('%Y-%m-%d') if sale[2] else 'N/A'
        })
    
    # Create DataFrame
    df = pd.DataFrame(data)
    
    # Debug: Print column names to check
    print("DataFrame columns:", df.columns.tolist())
    print("DataFrame shape:", df.shape)
    
    # Create summary statistics with error handling
    if df.empty:
        summary_data = {
            'Metric': ['Total Revenue', 'Total Cost', 'Gross Profit', 'Profit Margin (%)', 'Total Sales'],
            'Amount': [0, 0, 0, 0, 0]
        }
    else:
        summary_data = {
            'Metric': ['Total Revenue', 'Total Cost', 'Gross Profit', 'Profit Margin (%)', 'Total Sales'],
            'Amount': [
                df['Revenue (KES)'].sum(),
                df['Cost (KES)'].sum(),
                df['Profit (KES)'].sum(),
                (df['Profit (KES)'].sum() / df['Revenue (KES)'].sum() * 100) if df['Revenue (KES)'].sum() > 0 else 0,
                len(df)
            ]
        }
    summary_df = pd.DataFrame(summary_data)
    
    # Create Excel file in memory
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Write sales data
        df.to_excel(writer, sheet_name='Sales Details', index=False)
        
        # Write summary data
        summary_df.to_excel(writer, sheet_name='Summary', index=False)
        
        # Format sales worksheet
        workbook = writer.book
        worksheet_sales = writer.sheets['Sales Details']
        worksheet_summary = writer.sheets['Summary']
        
        # Import openpyxl styles
        from openpyxl.styles import Font, PatternFill, Border, Side
        
        # Define styles
        header_font_sales = Font(bold=True, color='FFFFFF')
        header_fill_sales = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        
        header_font_summary = Font(bold=True, color='FFFFFF')
        header_fill_summary = PatternFill(start_color='28a745', end_color='28a745', fill_type='solid')
        
        # Apply formatting to sales headers
        for col_num, header in enumerate(df.columns, 1):
            cell = worksheet_sales.cell(row=1, column=col_num)
            cell.font = header_font_sales
            cell.fill = header_fill_sales
        
        # Apply formatting to summary headers
        for col_num, header in enumerate(summary_df.columns, 1):
            cell = worksheet_summary.cell(row=1, column=col_num)
            cell.font = header_font_summary
            cell.fill = header_fill_summary
        
        # Auto-adjust column widths for both sheets
        for worksheet in [worksheet_sales, worksheet_summary]:
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
    
    # Create response
    output.seek(0)
    response = Response(
        output.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={
            'Content-Disposition': f'attachment; filename=financial_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        }
    )
    
    return response


@app.route('/admin/reports/pdf')
@login_required(UserRole.ADMIN)
def generate_financial_pdf():
    """Generate financial reports as PDF"""
    from flask import make_response
    from datetime import datetime
    
    # Get financial data
    total_revenue = (
        db.session.query(
            func.coalesce(
                func.sum(Sale.quantity * Product.selling_price), 0
            )
        )
        .join(Product, Product.id == Sale.product_id)
        .scalar()
    )
    
    total_cost = (
        db.session.query(
            func.coalesce(
                func.sum(Sale.quantity * Product.buying_price), 0
            )
        )
        .join(Product, Product.id == Sale.product_id)
        .scalar()
    )
    
    gross_profit = total_revenue - total_cost
    
    latest_sales = (
        db.session.query(Sale, Product)
        .join(Product, Product.id == Sale.product_id)
        .order_by(Sale.created_at.desc())
        .limit(10)
        .all()
    )
    
    # Generate HTML for PDF
    html_content = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <title>Financial Report - Piston Automotive Garage</title>
        <style>
            body {{ font-family: Arial, sans-serif; margin: 20px; }}
            .header {{ text-align: center; margin-bottom: 30px; }}
            .summary {{ background: #f8f9fa; padding: 20px; border-radius: 10px; margin-bottom: 30px; }}
            .summary-item {{ display: flex; justify-content: space-between; margin-bottom: 10px; }}
            .sales-table {{ width: 100%; border-collapse: collapse; margin-top: 20px; }}
            .sales-table th, .sales-table td {{ border: 1px solid #ddd; padding: 12px; text-align: left; }}
            .sales-table th {{ background-color: #f2f2f2; font-weight: bold; }}
            .footer {{ margin-top: 30px; text-align: center; font-size: 12px; color: #666; }}
        </style>
    </head>
    <body>
        <div class="header">
            <h1>Financial Report</h1>
            <h2>Piston Automotive Garage</h2>
            <p>Generated on: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}</p>
        </div>
        
        <div class="summary">
            <h3>Financial Summary</h3>
            <div class="summary-item">
                <strong>Total Revenue:</strong>
                <span>KES {total_revenue:,.2f}</span>
            </div>
            <div class="summary-item">
                <strong>Total Cost:</strong>
                <span>KES {total_cost:,.2f}</span>
            </div>
            <div class="summary-item">
                <strong>Gross Profit:</strong>
                <span>KES {gross_profit:,.2f}</span>
            </div>
            <div class="summary-item">
                <strong>Profit Margin:</strong>
                <span>{f"{(gross_profit/total_revenue*100):.1f}%" if total_revenue > 0 else "0%"}</span>
            </div>
        </div>
        
        <h3>Recent Sales</h3>
        <table class="sales-table">
            <thead>
                <tr>
                    <th>Sale ID</th>
                    <th>Product</th>
                    <th>Quantity</th>
                    <th>Revenue</th>
                    <th>Date</th>
                </tr>
            </thead>
            <tbody>
    """
    
    for sale, product in latest_sales:
        revenue = sale.quantity * product.selling_price
        html_content += f"""
                <tr>
                    <td>{sale.id}</td>
                    <td>{product.name}</td>
                    <td>{sale.quantity}</td>
                    <td>KES {revenue:,.2f}</td>
                    <td>{sale.created_at.strftime('%Y-%m-%d')}</td>
                </tr>
        """
    
    html_content += """
            </tbody>
        </table>
        
        <div class="footer">
            <p>This is an automatically generated financial report from Piston Automotive Garage Management System.</p>
        </div>
    </body>
    </html>
    """
    
    # For now, return the HTML as a response (in production, you'd use a PDF library like WeasyPrint or ReportLab)
    response = make_response(html_content)
    response.headers['Content-Type'] = 'text/html'
    response.headers['Content-Disposition'] = f'attachment; filename=financial_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.html'
    
    return response


@app.route('/admin/invoicing')
@login_required(UserRole.ADMIN)
def admin_invoicing():
    """Admin invoicing page - manage customer invoices"""
    invoices = Invoice.query.order_by(Invoice.created_at.desc()).all()
    return render_template('admin_invoicing.html', invoices=invoices)


@app.route('/admin/invoices/export')
@login_required(UserRole.ADMIN)
def export_invoices():
    """Export invoices as Excel file"""
    import pandas as pd
    from io import BytesIO
    from flask import Response
    
    # Get all invoices with their items
    invoices = Invoice.query.order_by(Invoice.created_at.desc()).all()
    
    invoice_data = []
    for invoice in invoices:
        # Get invoice items
        items = InvoiceItem.query.filter_by(invoice_id=invoice.id).all()
        
        for item in items:
            invoice_data.append({
                'Invoice Number': invoice.invoice_number,
                'Customer Name': invoice.customer_name,
                'Customer Email': invoice.customer_email,
                'Customer Phone': invoice.customer_phone,
                'Issue Date': invoice.issue_date.strftime('%Y-%m-%d'),
                'Due Date': invoice.due_date.strftime('%Y-%m-%d'),
                'Status': invoice.status.value,
                'Item Description': item.description,
                'Quantity': item.quantity,
                'Unit Price': item.unit_price,
                'Line Total': item.line_total,
                'Subtotal': invoice.subtotal,
                'Tax Rate (%)': invoice.tax_rate,
                'Tax Amount': invoice.tax_amount,
                'Total Amount': invoice.total_amount,
                'Notes': invoice.notes or '',
                'Created At': invoice.created_at.strftime('%Y-%m-%d %H:%M:%S')
            })
        
        # If no items, still add the invoice header
        if not items:
            invoice_data.append({
                'Invoice Number': invoice.invoice_number,
                'Customer Name': invoice.customer_name,
                'Customer Email': invoice.customer_email,
                'Customer Phone': invoice.customer_phone,
                'Issue Date': invoice.issue_date.strftime('%Y-%m-%d'),
                'Due Date': invoice.due_date.strftime('%Y-%m-%d'),
                'Status': invoice.status.value,
                'Item Description': '',
                'Quantity': 0,
                'Unit Price': 0,
                'Line Total': 0,
                'Subtotal': invoice.subtotal,
                'Tax Rate (%)': invoice.tax_rate,
                'Tax Amount': invoice.tax_amount,
                'Total Amount': invoice.total_amount,
                'Notes': invoice.notes or '',
                'Created At': invoice.created_at.strftime('%Y-%m-%d %H:%M:%S')
            })
    
    # Create DataFrame
    df = pd.DataFrame(invoice_data)
    
    # Create Excel file in memory
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Write main data
        df.to_excel(writer, sheet_name='Invoices', index=False)
        
        # Get workbook and worksheet for formatting
        workbook = writer.book
        worksheet = writer.sheets['Invoices']
        
        # Format headers
        from openpyxl.styles import Font, PatternFill, Border, Side
        
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                            top=Side(style='thin'), bottom=Side(style='thin'))
        
        # Apply formatting to header row
        for col_num, header in enumerate(df.columns, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.style = header_font
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    # Create response
    output.seek(0)
    response = Response(
        output.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={
            'Content-Disposition': f'attachment; filename=invoices_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        }
    )
    
    return response


@app.route('/admin/invoice/create', methods=['GET', 'POST'])
@login_required(UserRole.ADMIN)
def create_invoice():
    """Create new invoice"""
    if request.method == 'POST':
        try:
            # Get form data
            customer_name = request.form.get('customer_name')
            customer_email = request.form.get('customer_email')
            customer_phone = request.form.get('customer_phone')
            issue_date_str = request.form.get('issue_date')
            due_date_str = request.form.get('due_date')
            notes = request.form.get('notes')
            
            # Parse dates
            issue_date = datetime.strptime(issue_date_str, '%Y-%m-%d') if issue_date_str else datetime.utcnow()
            due_date = datetime.strptime(due_date_str, '%Y-%m-%d') if due_date_str else datetime.utcnow() + timedelta(days=30)
            
            # Get invoice items
            items_data = request.form.getlist('items[]')
            
            # Calculate totals
            subtotal = 0.0
            tax_rate = 16.0
            
            # Create invoice
            invoice = Invoice(
                invoice_number=f"INV-{datetime.now().strftime('%Y%m%d')}-{Invoice.query.count() + 1:04d}",
                customer_name=customer_name,
                customer_email=customer_email,
                customer_phone=customer_phone,
                issue_date=issue_date,
                due_date=due_date,
                status=InvoiceStatus.DRAFT,
                subtotal=subtotal,
                tax_rate=tax_rate,
                tax_amount=subtotal * (tax_rate / 100),
                total_amount=subtotal * (1 + tax_rate / 100),
                notes=notes
            )
            
            db.session.add(invoice)
            db.session.flush()  # Get the ID without committing
            
            # Add invoice items
            for item_data in items_data:
                if item_data:
                    parts = item_data.split('|')
                    if len(parts) >= 3:
                        product_id = int(parts[0])
                        description = parts[1]
                        quantity = int(parts[2])
                        
                        # Get product price
                        product = Product.query.get(product_id)
                        unit_price = product.selling_price if product else 0
                        line_total = unit_price * unit_price
                        
                        subtotal += line_total
                        
                        # Create invoice item
                        invoice_item = InvoiceItem(
                            invoice_id=invoice.id,
                            product_id=product_id,
                            description=description,
                            quantity=quantity,
                            unit_price=unit_price,
                            line_total=line_total
                        )
                        db.session.add(invoice_item)
            
            # Update invoice totals
            invoice.subtotal = subtotal
            invoice.tax_amount = subtotal * (tax_rate / 100)
            invoice.total_amount = subtotal * (1 + tax_rate / 100)
            
            db.session.commit()
            
            flash(f'Invoice {invoice.invoice_number} created successfully!', 'success')
            return redirect(url_for('admin_invoicing'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error creating invoice: {str(e)}', 'danger')
            return redirect(url_for('admin_invoicing'))
    
    # GET request - show create form
    products = Product.query.all()
    return render_template('create_invoice.html', products=products)





@app.route('/setup-admin', methods=['GET', 'POST'])
def setup_admin():
    """First-time admin setup - only accessible when no admins exist"""
    print("DEBUG: Setup admin route accessed")
    
    # Check if any admin already exists
    admin_count = User.query.filter_by(role=UserRole.ADMIN).count()
    print(f"DEBUG: Current admin count: {admin_count}")
    
    if admin_count > 0:
        flash('System already has admin users. Please login.', 'info')
        return redirect(url_for('login'))
    
    # Handle form submission manually for setup
    if request.method == 'POST':
        print(f"DEBUG: Form method: {request.method}")
        
        username = request.form.get('username')
        email = request.form.get('email')
        password = request.form.get('password')
        confirm_password = request.form.get('confirm_password')
        
        print(f"DEBUG: Form data - Username: {username}")
        print(f"DEBUG: Form data - Email: {email}")
        print(f"DEBUG: Form data - Password provided: {'Yes' if password else 'No'}")
        
        # Basic validation
        errors = []
        if not username:
            errors.append('Username is required')
        if not email:
            errors.append('Email is required')
        if not password:
            errors.append('Password is required')
        if password != confirm_password:
            errors.append('Passwords do not match')
        
        if errors:
            for error in errors:
                flash(f'❌ {error}', 'danger')
            return render_template('setup_admin.html', form=None, is_setup=True)
        
        # Check if user already exists
        existing_user = User.query.filter((User.username == username) | (User.email == email)).first()
        if existing_user:
            flash('❌ Username or email already exists!', 'danger')
            return render_template('setup_admin.html', form=None, is_setup=True)

        hashed_password = generate_password_hash(password)

        first_admin = User(
            username=username,
            email=email,
            password=hashed_password,
            role=UserRole.ADMIN
        )

        try:
            db.session.add(first_admin)
            db.session.commit()
            
            print(f"DEBUG: Admin {username} created successfully!")
            print(f"DEBUG: Email: {email}")
            
            # Multiple success notifications
            flash(f'🎉 Admin account "{username}" created successfully!', 'success')
            flash(f'📧 Email: {email}', 'info')
            flash('🔐 Your admin account is now ready to use!', 'success')
            flash('📋 Next Step: Login below to access your admin dashboard', 'info')
            flash('🚀 After login, you will see the Admin button in navigation', 'warning')
            
            print("DEBUG: Flash messages set, redirecting to login...")
            return redirect(url_for('login'))

        except IntegrityError as e:
            db.session.rollback()
            if 'duplicate key value violates unique constraint' in str(e):
                flash('❌ Username or email already exists!', 'danger')
            else:
                app.logger.error(f'Error creating admin: {str(e)}')
                flash('❌ Error creating admin. Please try again.', 'danger')

    return render_template('setup_admin.html', form=None, is_setup=True)


@app.route('/admin/register', methods=['GET', 'POST'])
@login_required(UserRole.ADMIN)
def admin_register():
    """Admin-only route to register new admin users"""
    form = RegisterForm()

    if form.validate_on_submit():
        username = form.username.data
        email = form.email.data
        password = form.password.data

        hashed_password = generate_password_hash(password)

        new_admin = User(
            username=username,
            email=email,
            password=hashed_password,
            role=UserRole.ADMIN
        )

        try:
            db.session.add(new_admin)
            db.session.commit()
            flash(f'Admin user {username} created successfully!', 'success')
            return redirect(url_for('admin_dashboard'))

        except IntegrityError as e:
            db.session.rollback()
            if 'duplicate key value violates unique constraint' in str(e):
                flash('Username or email already exists!', 'danger')
            else:
                app.logger.error(f'Error creating admin: {str(e)}')
                flash('Error creating admin. Please try again.', 'danger')

    return render_template('admin_register.html', form=form)


@app.route('/debug-session')
def debug_session():
    """Debug route to check session data"""
    return {
        'username': session.get('username'),
        'user_role': session.get('user_role'),
        'session_data': dict(session)
    }

@app.route('/debug-users')
def debug_users():
    """Debug route to check all users in database"""
    users = User.query.all()
    user_list = []
    for user in users:
        user_list.append({
            'id': user.id,
            'username': user.username,
            'email': user.email,
            'role': user.role.value
        })
    return {'users': user_list, 'total_count': len(user_list)}


@app.route('/logout')
def logout():
    # Clear all session data
    session.clear()
    flash('Logout successful!', 'success')
    return redirect(url_for('landing'))



if __name__ == '__main__':
    with app.app_context():
        db.create_all()
        
        # Check if any admin exists
        admin_count = User.query.filter_by(role=UserRole.ADMIN).count()
        
        if admin_count == 0:
            print("🚨 No admin users found!")
            print("📝 Please register an admin user first.")
            print("🔑 Navigate to: http://localhost:5500/setup-admin")
            print("🔑 This will be your only access to create the first admin account.")
        else:
            print(f"✅ {admin_count} admin user(s) found. System ready.")
    
    app.run(debug=True, host='127.0.0.1', port=5500, use_reloader=False)
